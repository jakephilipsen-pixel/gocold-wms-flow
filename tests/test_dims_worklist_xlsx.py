from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from analysis.dims_worklist import build_worklist, WORKLIST_COLUMNS
from analysis.dims_worklist_xlsx import write_worklist_xlsx

_YELLOW = "FFFFE0"
_GREY = "D9D9D9"


def _dims(rows):
    cols = ["product_code", "outer_l_mm", "outer_w_mm", "outer_h_mm",
            "outer_weight_kg", "inner_pack_qty"]
    return pd.DataFrame([{c: r.get(c) for c in cols} for r in rows], columns=cols)


def _prod(code, name, uoms):
    return {"id": f"id-{code}", "references": {"code": code},
            "name": name, "unitOfMeasures": uoms}


def _col_letter(name):
    return WORKLIST_COLUMNS.index(name) + 1


def _build(tmp_path) -> Path:
    dims = _dims([
        {"product_code": "INNER1", "outer_l_mm": 50, "outer_w_mm": 40, "outer_h_mm": 30,
         "outer_weight_kg": 0.2, "inner_pack_qty": 1},
        {"product_code": "CART12", "outer_l_mm": 300, "outer_w_mm": 200, "outer_h_mm": 150,
         "outer_weight_kg": 6.0, "inner_pack_qty": 12},
        {"product_code": "NOWT", "outer_l_mm": 10, "outer_w_mm": 10, "outer_h_mm": 10,
         "outer_weight_kg": float("nan"), "inner_pack_qty": 1},
    ])
    prods = [
        _prod("INNER1", "an inner", {"EA": {"baseQty": 1}}),
        _prod("CART12", "a carton", {"EA": {"baseQty": 1}, "CT": {"baseQty": 12}}),
        _prod("NOWT", "no weight", {"EA": {"baseQty": 1}}),
    ]
    wl = build_worklist(dims, prods)
    out = tmp_path / "worklist.xlsx"
    write_worklist_xlsx(wl, out)
    return out


def test_writes_workbook_with_headers(tmp_path):
    out = _build(tmp_path)
    assert out.exists()
    ws = load_workbook(out).active
    # row 1 is exactly the worklist header; data starts at row 2
    first_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert first_row == WORKLIST_COLUMNS


def _fill_hex(cell):
    rgb = cell.fill.fgColor.rgb
    return None if rgb is None else str(rgb)[-6:]


def test_carton_each_cells_highlighted_yellow(tmp_path):
    out = _build(tmp_path)
    ws = load_workbook(out).active
    # data rows start at row 2 (row 1 = header). CART12 is the 2nd data row → row 3.
    r = 3
    for col_name in ("each_l_mm", "each_w_mm", "each_h_mm"):
        cell = ws.cell(row=r, column=_col_letter(col_name))
        assert _fill_hex(cell) == _YELLOW


def test_inner_each_cells_greyed_with_captured_marker(tmp_path):
    out = _build(tmp_path)
    ws = load_workbook(out).active
    r = 2  # INNER1 = first data row
    cell = ws.cell(row=r, column=_col_letter("each_l_mm"))
    assert _fill_hex(cell) == _GREY
    assert cell.value == "= captured"


def test_weight_pending_cell_highlighted(tmp_path):
    out = _build(tmp_path)
    ws = load_workbook(out).active
    r = 4  # NOWT = third data row
    cell = ws.cell(row=r, column=_col_letter("outer_weight_kg"))
    assert _fill_hex(cell) == _YELLOW
