"""Render a dims worklist DataFrame to a highlighted xlsx.

Row 1 is the header (= WORKLIST_COLUMNS); data starts at row 2. Cells that
need operator action are filled yellow; inner rows' each-cells are greyed and
marked "= captured" so they are visibly not-for-fill.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analysis.dims_worklist import WORKLIST_COLUMNS

_HEADER_BG = "1F2937"      # Go Cold dark
_YELLOW = "FFFFE0"          # needs fill / resolve
_GREY = "D9D9D9"            # not-for-fill (inner each cells)
_EACH_COLS = ("each_l_mm", "each_w_mm", "each_h_mm")

_yellow_fill = PatternFill("solid", fgColor=f"FF{_YELLOW}")
_grey_fill = PatternFill("solid", fgColor=f"FF{_GREY}")
_header_fill = PatternFill("solid", fgColor=f"FF{_HEADER_BG}")


def _col_idx(name: str) -> int:
    return WORKLIST_COLUMNS.index(name) + 1


def _cell_value(v: object):
    """openpyxl can't write pandas NA/NaN — coerce to None."""
    if v is None or (isinstance(v, float) and pd.isna(v)) or v is pd.NA:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def write_worklist_xlsx(df: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dims Worklist"

    # header
    for j, name in enumerate(WORKLIST_COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=name)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = _header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = max(12, len(name) + 2)

    # data
    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 2
        for name in WORKLIST_COLUMNS:
            ws.cell(row=r, column=_col_idx(name), value=_cell_value(row[name]))

        kind = row["kind"]
        # each cells
        if kind == "carton":
            for name in _EACH_COLS:
                ws.cell(row=r, column=_col_idx(name)).fill = _yellow_fill
        elif kind == "inner":
            for name in _EACH_COLS:
                cell = ws.cell(row=r, column=_col_idx(name))
                cell.value = "= captured"
                cell.fill = _grey_fill
        else:  # unknown → resolve ipq + measure each
            ws.cell(row=r, column=_col_idx("inner_pack_qty")).fill = _yellow_fill
            for name in _EACH_COLS:
                ws.cell(row=r, column=_col_idx(name)).fill = _yellow_fill

        # weight gap
        if bool(row["weight_pending"]):
            ws.cell(row=r, column=_col_idx("outer_weight_kg")).fill = _yellow_fill

    ws.freeze_panes = "C2"  # freeze header row + code/name columns
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
