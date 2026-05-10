"""Generate the carton-dimension capture template.

Outputs an Excel workbook with one row per SKU, sorted by measurement
priority (units/day desc), with empty columns for the warehouse team to
fill in. When complete, this gets either re-imported to CC or fed back
into our local product master for the slotting recommendations.

Sheet layout deliberately compact and human-friendly — designed to be
filled in on a phone or tablet on the warehouse floor.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Go Cold brand colours (per CLAUDE.md instructions)
GC_GREEN = "00C452"
GC_BLUE = "0096CC"
GC_DARK = "003366"
GC_MID = "0076A8"


def _write_header_row(ws, row: int, columns: list[tuple[str, int]]) -> None:
    """Write a header row with Go Cold styling."""
    for col_idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=GC_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 36


def write_capture_template(
    sku_metrics: pd.DataFrame, out_path: Path
) -> None:
    """Write the SKU dimension capture workbook to out_path."""
    wb = Workbook()

    # --- main sheet: capture
    ws = wb.active
    ws.title = "SKU Capture"

    # Title row
    ws.cell(row=1, column=1, value="Go Cold — Forage SKU Dimension Capture").font = Font(
        name="Arial", bold=True, size=14, color=GC_DARK
    )
    ws.cell(row=2, column=1, value=(
        "Measure each SKU at OUTER carton level. Length × Width × Height in mm; "
        "Weight in kg (full carton). Inner pack qty = units inside one outer carton."
    )).font = Font(name="Arial", size=10, italic=True)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=10)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=10)

    columns = [
        ("Priority", 9),
        ("Product Code", 14),
        ("Product Name", 38),
        ("Type", 11),
        ("Units/day", 11),
        ("ABC", 7),
        ("Outer L (mm)", 12),
        ("Outer W (mm)", 12),
        ("Outer H (mm)", 12),
        ("Outer Weight (kg)", 14),
        ("Inner Pack Qty", 13),
        ("Outer Carton Qty", 14),
        ("Pallet TI x HI", 13),
        ("Notes", 30),
        ("Measured By", 14),
        ("Date Measured", 14),
    ]
    _write_header_row(ws, row=4, columns=columns)

    # data rows
    df = sku_metrics.reset_index()  # 'product_code' becomes a column
    if "product_code" not in df.columns:
        df = df.rename(columns={df.columns[0]: "product_code"})

    df = df.sort_values("measure_priority")

    start_row = 5
    abc_fills = {
        "A": PatternFill("solid", fgColor=GC_GREEN),
        "B": PatternFill("solid", fgColor=GC_BLUE),
        "C": PatternFill("solid", fgColor="CCCCCC"),
    }
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + i
        ws.cell(row=r, column=1, value=int(row.get("measure_priority", i + 1)))
        ws.cell(row=r, column=2, value=row.get("product_code", "")).font = Font(name="Arial", bold=True)
        ws.cell(row=r, column=3, value=row.get("product_name", ""))
        ws.cell(row=r, column=4, value=row.get("type", ""))
        ws.cell(row=r, column=5, value=round(float(row.get("units_per_day", 0) or 0), 1))
        abc = row.get("abc_class", "C") or "C"
        c = ws.cell(row=r, column=6, value=abc)
        c.alignment = Alignment(horizontal="center")
        if abc in abc_fills:
            c.fill = abc_fills[abc]
            c.font = Font(name="Arial", bold=True,
                          color="FFFFFF" if abc != "C" else "000000")
        # empty input columns 7-13
        for col in range(7, 14):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFE0")
        # notes / measured by / date — also empty, white
        for col in range(14, 17):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFFF")

    ws.freeze_panes = "C5"  # freeze priority + code columns + header rows
    ws.auto_filter.ref = f"A4:{get_column_letter(len(columns))}{start_row + len(df) - 1}"

    # --- legend sheet
    legend = wb.create_sheet("Legend")
    legend.cell(row=1, column=1, value="How to use this sheet").font = Font(
        name="Arial", bold=True, size=14, color=GC_DARK
    )
    rows = [
        "",
        "1. Work top-down (Priority column). The top SKUs ship more "
        "and unblock the slotting analysis fastest.",
        "",
        "2. For each SKU, measure the OUTER carton (the one delivered on a pallet):",
        "   - L × W × H in millimetres",
        "   - Weight in kilograms (full carton)",
        "",
        "3. Inner Pack Qty = number of UNITS inside ONE outer carton.",
        "   E.g. a carton of 24 cans of soda → Inner Pack Qty = 24",
        "",
        "4. Outer Carton Qty = how many SAP/CC counts as 1 picking unit.",
        "   For most SKUs this matches Inner Pack Qty (you pick whole cartons).",
        "   For SKUs picked as inners, this is the inner count.",
        "",
        "5. Pallet TI × HI = how the SKU stacks on a pallet (e.g. 6x5).",
        "   Skip if you don't know — we can derive from cube later.",
        "",
        "6. Initial of who measured + date. If you're unsure of any number, "
        "leave the cell empty and add a note rather than guessing.",
        "",
        "ABC class colour key:",
        "  A (green) = top 80% of volume — measure these first",
        "  B (blue)  = next 15% of volume",
        "  C (grey)  = remaining 5% — lowest priority",
    ]
    for i, text in enumerate(rows, start=2):
        legend.cell(row=i, column=1, value=text).font = Font(name="Arial", size=11)
    legend.column_dimensions["A"].width = 100

    wb.save(out_path)
