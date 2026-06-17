"""Trim a dims worklist down to a print-friendly measuring sheet.

The full worklist has one row per CC product (~461 for Forage), most of which
are already complete and not actionable. For a pen-and-paper measuring walk we
want only the rows that need something, split into non-overlapping job groups:

  - full_capture : kind == "unknown" — no usable local dims (new SKUs in CC, or
                   a blank inner-pack-qty). Need outer L/W/H + weight + ipq.
  - measure_each : kind == "carton" — captured outer dims exist, but the
                   each-level L/W/H were never measured. Need each L/W/H (and a
                   weight if that's also pending).
  - weigh_only   : kind == "inner" AND weight pending — dims are complete, just
                   the weight is missing. Need a scale.

Complete inner rows (dims + weight present) are dropped entirely.

`partition_measuring` is pure; styling/IO lives in `write_measuring_sheet`.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Group order = the order sections appear in the sheet.
GROUP_ORDER = ("measure_each", "full_capture", "weigh_only")

GROUP_TITLES = {
    "measure_each": "MEASURE EACH UNIT — L x W x H (mm) of one inner unit",
    "full_capture": "FULL CAPTURE — new/unknown SKU: outer L/W/H, weight, inner-pack-qty",
    "weigh_only": "WEIGH ONLY — dims already captured, weight missing (kg)",
}

# Per-group column layout: (worklist column OR None for a blank write-in box,
# printed header). None entries render as empty bordered boxes to fill by hand.
_GROUP_COLUMNS = {
    "measure_each": [
        ("product_code", "Product Code"),
        ("product_name", "Product Name"),
        ("carton_uom_code", "Carton UoM"),
        ("inner_pack_qty", "Inners/Outer"),
        (None, "Each L (mm)"),
        (None, "Each W (mm)"),
        (None, "Each H (mm)"),
        ("__weight_box__", "Weight (kg)"),
    ],
    "full_capture": [
        ("product_code", "Product Code"),
        ("product_name", "Product Name"),
        (None, "Outer L (mm)"),
        (None, "Outer W (mm)"),
        (None, "Outer H (mm)"),
        (None, "Weight (kg)"),
        (None, "Inners/Outer"),
    ],
    "weigh_only": [
        ("product_code", "Product Code"),
        ("product_name", "Product Name"),
        ("outer_l_mm", "L (mm)"),
        ("outer_w_mm", "W (mm)"),
        ("outer_h_mm", "H (mm)"),
        (None, "Weight (kg)"),
    ],
}


def partition_measuring(wl: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Split a worklist into the three non-overlapping measuring groups.

    Each actionable row lands in exactly one group; complete inner rows are
    excluded. Returns a dict keyed by GROUP_ORDER, each a DataFrame sorted by
    product_code.
    """
    weight_pending = wl["weight_pending"].astype(bool)
    groups = {
        "measure_each": wl[wl["kind"] == "carton"],
        "full_capture": wl[wl["kind"] == "unknown"],
        "weigh_only": wl[(wl["kind"] == "inner") & weight_pending],
    }
    return {
        k: g.sort_values("product_code").reset_index(drop=True)
        for k, g in groups.items()
    }


def _cell_value(v: object):
    """openpyxl can't write pandas NA/NaN — coerce to None."""
    try:
        return None if pd.isna(v) else v
    except (TypeError, ValueError):
        return v


_HEADER_BG = "FF1F2937"
_SECTION_BG = "FFB45309"   # amber section banner
_BOX_BG = "FFFFFFE0"        # write-in boxes (pale yellow)
_thin = Side(style="thin", color="FF999999")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def write_measuring_sheet(groups: dict[str, pd.DataFrame], path: Path) -> None:
    """Render the measuring groups to a single print-friendly xlsx sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Measuring Sheet"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    r = 1
    for key in GROUP_ORDER:
        df = groups.get(key)
        if df is None or df.empty:
            continue
        layout = _GROUP_COLUMNS[key]
        ncols = len(layout)

        # section banner
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        banner = ws.cell(row=r, column=1, value=f"{GROUP_TITLES[key]}   ({len(df)} SKUs)")
        banner.font = Font(name="Arial", bold=True, color="FFFFFFFF", size=12)
        banner.fill = PatternFill("solid", fgColor=_SECTION_BG)
        banner.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        # column headers
        header_row = r
        for j, (_, title) in enumerate(layout, start=1):
            c = ws.cell(row=r, column=j, value=title)
            c.font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=_HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border
        ws.row_dimensions[r].height = 28
        r += 1

        # data rows
        for _, row in df.iterrows():
            for j, (src, _) in enumerate(layout, start=1):
                cell = ws.cell(row=r, column=j)
                cell.border = _border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if src is None:
                    cell.fill = PatternFill("solid", fgColor=_BOX_BG)
                elif src == "__weight_box__":
                    # blank box when pending, else show the known weight
                    if bool(row.get("weight_pending")):
                        cell.fill = PatternFill("solid", fgColor=_BOX_BG)
                    else:
                        cell.value = _cell_value(row.get("outer_weight_kg"))
                else:
                    cell.value = _cell_value(row.get(src))
                    if src == "product_name":
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        # repeat the group's header row across pages while in this section
        r += 1  # spacer before next section
        ws.row_breaks  # (page break handled by fitToWidth; spacer keeps sections distinct)

    # column widths: identity wide, boxes comfortable for handwriting
    widths = [14, 34, 11, 12, 12, 12, 12, 12]
    for j in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(j)].width = (
            widths[j - 1] if j - 1 < len(widths) else 12
        )

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
